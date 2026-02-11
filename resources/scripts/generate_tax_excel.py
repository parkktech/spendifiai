#!/usr/bin/env python3
"""
Generate a professional multi-tab Excel tax workbook from SpendWise data.
Usage: python3 generate_tax_excel.py <input.json> <output.xlsx>
"""
import json, sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

def load_data(path):
    with open(path) as f:
        return json.load(f)

def style_header_row(ws, row, cols, fill_color="1a5276"):
    fill = PatternFill("solid", fgColor=fill_color)
    font = Font(bold=True, color="FFFFFF", size=11)
    border = Side(style="thin", color="cccccc")
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=Side(style="medium", color="333333"))

def currency_fmt(ws, cell_ref):
    ws[cell_ref].number_format = '$#,##0.00'

def pct_fmt(ws, cell_ref):
    ws[cell_ref].number_format = '0.0%'

def auto_width(ws, min_w=10, max_w=45):
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            val = str(cell.value) if cell.value else ""
            max_len = max(max_len, len(val))
        ws.column_dimensions[col_letter].width = min(max(max_len + 3, min_w), max_w)

def create_summary_sheet(wb, data):
    ws = wb.active
    ws.title = "Tax Summary"
    ws.sheet_properties.tabColor = "1a5276"

    s = data["summary"]
    p = data["profile"]
    u = data["user"]

    # Title
    ws.merge_cells("A1:F1")
    ws["A1"] = f"SpendWise Tax Export — {data['year']}"
    ws["A1"].font = Font(bold=True, size=18, color="1a5276")
    ws["A1"].alignment = Alignment(horizontal="left")

    ws.merge_cells("A2:F2")
    ws["A2"] = f"Prepared for {u['name']} ({u['email']}) on {s['generated_at'][:10]}"
    ws["A2"].font = Font(size=10, color="666666")

    # Profile section
    row = 4
    ws.cell(row=row, column=1, value="TAXPAYER PROFILE").font = Font(bold=True, size=12, color="1a5276")
    row += 1
    profile_fields = [
        ("Employment Type", p.get("employment_type", "—")),
        ("Business Type", p.get("business_type", "—")),
        ("Filing Status", p.get("filing_status", "—")),
        ("Est. Tax Bracket", f"{p.get('tax_bracket', 22)}%"),
        ("Home Office", "Yes" if p.get("has_home_office") else "No"),
    ]
    for label, val in profile_fields:
        ws.cell(row=row, column=1, value=label).font = Font(bold=True, size=10)
        ws.cell(row=row, column=2, value=val)
        row += 1

    # Totals section
    row += 1
    ws.cell(row=row, column=1, value="DEDUCTION SUMMARY").font = Font(bold=True, size=12, color="1a5276")
    row += 1

    green_fill = PatternFill("solid", fgColor="e8f5e9")
    summary_rows = [
        ("Total Deductible (Bank Transactions)", s["total_deductible_transactions"]),
        ("Total Deductible (Email Order Items)", s["total_deductible_items"]),
        ("Grand Total Deductible", s["grand_total_deductible"]),
        ("Estimated Tax Savings", s["estimated_tax_savings"]),
    ]
    for label, val in summary_rows:
        ws.cell(row=row, column=1, value=label).font = Font(bold=True, size=11)
        c = ws.cell(row=row, column=2, value=val)
        c.number_format = '$#,##0.00'
        c.font = Font(bold=True, size=11)
        if "Grand Total" in label or "Tax Savings" in label:
            ws.cell(row=row, column=1).fill = green_fill
            c.fill = green_fill
            c.font = Font(bold=True, size=13, color="2e7d32")
        row += 1

    ws.cell(row=row, column=1, value=f"Based on {s['effective_rate_used']*100:.0f}% effective rate").font = Font(size=9, italic=True, color="999999")
    row += 2

    # Accounts used
    ws.cell(row=row, column=1, value="LINKED ACCOUNTS").font = Font(bold=True, size=12, color="1a5276")
    row += 1
    headers = ["Institution", "Account", "Type", "Last 4", "Purpose", "Business Name", "Entity Type", "EIN"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=row, column=i, value=h)
    style_header_row(ws, row, len(headers), "34495e")
    row += 1
    for acct in data.get("accounts", []):
        ws.cell(row=row, column=1, value=acct.get("institution", ""))
        ws.cell(row=row, column=2, value=acct.get("account", ""))
        ws.cell(row=row, column=3, value=acct.get("type", ""))
        ws.cell(row=row, column=4, value=acct.get("mask", ""))
        purpose_cell = ws.cell(row=row, column=5, value=acct.get("purpose", ""))
        if acct.get("purpose") == "business":
            purpose_cell.font = Font(bold=True, color="1a5276")
        ws.cell(row=row, column=6, value=acct.get("business_name", ""))
        ws.cell(row=row, column=7, value=acct.get("entity_type", ""))
        ws.cell(row=row, column=8, value=acct.get("ein", ""))
        row += 1

    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 25
    auto_width(ws, min_w=12)

def create_schedule_c_sheet(wb, data):
    ws = wb.create_sheet("Schedule C Mapping")
    ws.sheet_properties.tabColor = "2e7d32"

    ws.merge_cells("A1:E1")
    ws["A1"] = f"IRS Schedule C Line Mapping — {data['year']}"
    ws["A1"].font = Font(bold=True, size=14, color="2e7d32")

    ws["A2"] = "Maps your deductible expenses to IRS Schedule C (Form 1040) lines"
    ws["A2"].font = Font(size=10, italic=True, color="666666")

    headers = ["Schedule C Line", "Description", "Amount", "# Items", "Categories Included"]
    row = 4
    for i, h in enumerate(headers, 1):
        ws.cell(row=row, column=i, value=h)
    style_header_row(ws, row, len(headers), "2e7d32")

    row = 5
    grand_total = 0
    for line in data.get("schedule_c_mapping", []):
        ws.cell(row=row, column=1, value=f"Line {line['line']}").font = Font(bold=True)
        ws.cell(row=row, column=2, value=line["label"])
        amt_cell = ws.cell(row=row, column=3, value=line["total"])
        amt_cell.number_format = '$#,##0.00'
        amt_cell.font = Font(bold=True)

        total_items = sum(c.get("items", 0) for c in line.get("categories", []))
        ws.cell(row=row, column=4, value=total_items)

        cat_names = ", ".join(c["name"] for c in line.get("categories", []))
        ws.cell(row=row, column=5, value=cat_names)

        # Meals are 50% deductible — flag it
        if line["line"] == "24b":
            ws.cell(row=row, column=2).value += " ⚠️ 50% deductible"
            ws.cell(row=row, column=2).font = Font(italic=True, color="e65100")

        grand_total += line["total"]
        row += 1

    # Total row
    row += 1
    ws.cell(row=row, column=1, value="TOTAL").font = Font(bold=True, size=12)
    total_cell = ws.cell(row=row, column=3, value=grand_total)
    total_cell.number_format = '$#,##0.00'
    total_cell.font = Font(bold=True, size=12, color="2e7d32")
    total_cell.fill = PatternFill("solid", fgColor="e8f5e9")

    auto_width(ws)

def create_category_sheet(wb, data):
    ws = wb.create_sheet("Deductions by Category")
    ws.sheet_properties.tabColor = "1565c0"

    ws.merge_cells("A1:E1")
    ws["A1"] = f"Deductions by Category — {data['year']}"
    ws["A1"].font = Font(bold=True, size=14, color="1565c0")

    headers = ["Tax Category", "Total Amount", "# Transactions", "Date Range"]
    row = 3
    for i, h in enumerate(headers, 1):
        ws.cell(row=row, column=i, value=h)
    style_header_row(ws, row, len(headers), "1565c0")

    row = 4
    for cat in data.get("deductions_by_category", []):
        ws.cell(row=row, column=1, value=cat["tax_category"])
        amt = ws.cell(row=row, column=2, value=cat["total"])
        amt.number_format = '$#,##0.00'
        amt.font = Font(bold=True)
        ws.cell(row=row, column=3, value=cat["item_count"])
        date_range = f"{cat.get('first_date', '')} — {cat.get('last_date', '')}"
        ws.cell(row=row, column=4, value=date_range)
        row += 1

    # Grand total with formula
    row += 1
    ws.cell(row=row, column=1, value="TOTAL").font = Font(bold=True, size=12)
    ws.cell(row=row, column=2, value=f"=SUM(B4:B{row-2})")
    ws.cell(row=row, column=2).number_format = '$#,##0.00'
    ws.cell(row=row, column=2).font = Font(bold=True, size=12, color="1565c0")

    auto_width(ws)

def create_transactions_sheet(wb, data):
    ws = wb.create_sheet("All Deductible Transactions")
    ws.sheet_properties.tabColor = "e65100"

    ws.merge_cells("A1:L1")
    ws["A1"] = f"Complete Deductible Transaction Detail — {data['year']}"
    ws["A1"].font = Font(bold=True, size=14, color="e65100")

    ws["A2"] = f"{len(data['deductible_transactions'])} bank transactions + {len(data['deductible_items'])} itemized email receipts"
    ws["A2"].font = Font(size=10, italic=True, color="666666")

    headers = [
        "Date", "Merchant", "Description", "Amount", "Category",
        "Tax Category", "Type", "Account", "Account Purpose",
        "Business Entity", "Confidence", "Verified"
    ]
    row = 4
    for i, h in enumerate(headers, 1):
        ws.cell(row=row, column=i, value=h)
    style_header_row(ws, row, len(headers), "e65100")

    # Bank transactions
    row = 5
    for tx in data.get("deductible_transactions", []):
        ws.cell(row=row, column=1, value=tx["date"])
        ws.cell(row=row, column=2, value=tx["merchant"])
        ws.cell(row=row, column=3, value=tx.get("description", ""))
        amt = ws.cell(row=row, column=4, value=tx["amount"])
        amt.number_format = '$#,##0.00'
        ws.cell(row=row, column=5, value=tx.get("category", ""))
        ws.cell(row=row, column=6, value=tx.get("tax_category", ""))
        ws.cell(row=row, column=7, value=tx.get("expense_type", ""))
        ws.cell(row=row, column=8, value=tx.get("account", ""))
        ws.cell(row=row, column=9, value=tx.get("account_purpose", ""))
        ws.cell(row=row, column=10, value=tx.get("business_name", ""))

        conf = tx.get("confidence")
        if conf:
            c = ws.cell(row=row, column=11, value=conf)
            c.number_format = '0%'
        verified = ws.cell(row=row, column=12, value="✓ User" if tx.get("user_confirmed") else "AI")
        if tx.get("user_confirmed"):
            verified.font = Font(color="2e7d32", bold=True)
        row += 1

    # Email order items (separate section)
    if data.get("deductible_items"):
        row += 1
        ws.cell(row=row, column=1, value="ITEMIZED EMAIL RECEIPTS").font = Font(bold=True, size=11, color="e65100")
        row += 1
        for item in data["deductible_items"]:
            ws.cell(row=row, column=1, value=item["date"])
            ws.cell(row=row, column=2, value=item["merchant"])
            desc = f"{item['product']}"
            if item.get("order_number"):
                desc += f" (Order #{item['order_number']})"
            ws.cell(row=row, column=3, value=desc)
            amt = ws.cell(row=row, column=4, value=item["amount"])
            amt.number_format = '$#,##0.00'
            ws.cell(row=row, column=5, value=item.get("category", ""))
            ws.cell(row=row, column=6, value=item.get("tax_category", ""))
            ws.cell(row=row, column=7, value="business")
            ws.cell(row=row, column=8, value="Email Receipt")
            ws.cell(row=row, column=12, value="AI")
            row += 1

    # Grand total
    row += 1
    ws.cell(row=row, column=3, value="GRAND TOTAL").font = Font(bold=True, size=12)
    ws.cell(row=row, column=4, value=f"=SUM(D5:D{row-2})")
    ws.cell(row=row, column=4).number_format = '$#,##0.00'
    ws.cell(row=row, column=4).font = Font(bold=True, size=12, color="e65100")

    auto_width(ws)

def create_subscriptions_sheet(wb, data):
    ws = wb.create_sheet("Business Subscriptions")
    ws.sheet_properties.tabColor = "6a1b9a"

    ws["A1"] = f"Recurring Business Expenses — {data['year']}"
    ws["A1"].font = Font(bold=True, size=14, color="6a1b9a")

    headers = ["Service", "Monthly Cost", "Annual Cost", "Category", "Frequency"]
    row = 3
    for i, h in enumerate(headers, 1):
        ws.cell(row=row, column=i, value=h)
    style_header_row(ws, row, len(headers), "6a1b9a")

    row = 4
    for sub in data.get("business_subscriptions", []):
        ws.cell(row=row, column=1, value=sub["service"])
        ws.cell(row=row, column=2, value=sub["monthly_cost"]).number_format = '$#,##0.00'
        ws.cell(row=row, column=3, value=sub["annual_cost"]).number_format = '$#,##0.00'
        ws.cell(row=row, column=4, value=sub["category"])
        ws.cell(row=row, column=5, value=sub["frequency"])
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=row, column=2, value=f"=SUM(B4:B{row-2})").number_format = '$#,##0.00'
    ws.cell(row=row, column=3, value=f"=SUM(C4:C{row-2})").number_format = '$#,##0.00'

    auto_width(ws)

def main():
    if len(sys.argv) < 3:
        print("Usage: generate_tax_excel.py <input.json> <output.xlsx>")
        sys.exit(1)

    data = load_data(sys.argv[1])
    wb = Workbook()

    create_summary_sheet(wb, data)
    create_schedule_c_sheet(wb, data)
    create_category_sheet(wb, data)
    create_transactions_sheet(wb, data)
    create_subscriptions_sheet(wb, data)

    # Print settings for all sheets
    for ws in wb.worksheets:
        ws.print_options.horizontalCentered = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.orientation = "landscape"

    wb.save(sys.argv[2])
    print(f"Generated: {sys.argv[2]}")

if __name__ == "__main__":
    main()
